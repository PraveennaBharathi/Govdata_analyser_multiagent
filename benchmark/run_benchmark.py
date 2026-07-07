"""
GovData Analyser — Benchmark Runner
Submits 50 questions to /query/async, polls for results, writes Excel.

Usage:
    python3 run_benchmark.py
    python3 run_benchmark.py --model mistral-small   # force model
    python3 run_benchmark.py --start 1 --end 10      # run subset
"""

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from questions import QUESTIONS

API_BASE = "http://localhost:8000"
POLL_INTERVAL_S = 4
POLL_TIMEOUT_S = 300  # Ollama reasoning tasks can take 2-3min

# ── Column definitions ────────────────────────────────────────────────────────

COLUMNS = [
    ("Q#",              8),
    ("Category",       14),
    ("Question",       52),
    ("Expected Domain",15),
    ("Actual Domain",  15),
    ("Status",         12),
    ("Model Used",     20),
    ("Latency (s)",    12),
    ("Relevant?",      10),
    ("Summary (first 300 chars)", 55),
    ("Full Answer",    80),
    ("Insights",       60),
    ("Parsed Filters", 28),
    ("Records Loaded", 14),
    ("Error",          35),
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def submit_query(question: str, model: str = "auto") -> dict:
    r = requests.post(
        f"{API_BASE}/query/async",
        json={"query": question, "model": model},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


def poll_task(task_id: str) -> dict:
    deadline = time.time() + POLL_TIMEOUT_S
    while time.time() < deadline:
        r = requests.get(f"{API_BASE}/task/{task_id}", timeout=10)
        r.raise_for_status()
        data = r.json()
        state = data.get("state", "")
        if state in ("SUCCESS", "FAILURE"):
            return data
        time.sleep(POLL_INTERVAL_S)
    return {"state": "TIMEOUT", "error": f"No response within {POLL_TIMEOUT_S}s"}


def extract_fields(task_result: dict) -> dict:
    """Pull useful fields from the task result for Excel columns."""
    out = {
        "actual_domain":  "",
        "model_used":     "",
        "latency_s":      "",
        "summary":        "",
        "full_answer":    "",
        "insights":       "",
        "parsed_filters": "",
        "records_loaded": "",
        "error":          "",
    }

    if task_result.get("state") == "FAILURE":
        out["error"] = task_result.get("error", "unknown")
        return out

    result = task_result.get("result") or {}

    # model
    out["model_used"] = result.get("model_preference", "")

    # domain
    parsed_q = result.get("parsed_query") or {}
    out["actual_domain"] = parsed_q.get("domain", "")
    filters = parsed_q.get("filters") or {}
    out["parsed_filters"] = json.dumps(filters) if filters else ""

    # timing — sum node timings for actual processing time
    steps = result.get("workflow_steps") or []
    total_ms = sum(s.get("duration_ms", 0) for s in steps if isinstance(s, dict))
    out["latency_s"] = round(total_ms / 1000, 2) if total_ms else ""

    # answer text
    inner = result.get("result") or {}
    out["full_answer"] = inner.get("conversational_response", "")
    out["summary"] = out["full_answer"][:300] if out["full_answer"] else ""

    # insights
    analysis = inner.get("analysis") or {}
    raw_insights = analysis.get("insights") or []
    if raw_insights:
        out["insights"] = "\n".join(f"• {i}" for i in raw_insights[:5])

    # records
    extraction = inner.get("extraction_summary") or {}
    out["records_loaded"] = extraction.get("records_count", "")

    return out


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
DOMAIN_COLORS = {
    "housing":     "D6E4BC",
    "labour":      "BDD7EE",
    "cross_domain":"F4CCCC",
    "environment": "FCE5CD",
    "business":    "EAD1DC",
    "demographics":"D9D2E9",
    "":            "F3F3F3",
}

def style_header(ws):
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def write_row(ws, row_num: int, q_id, category, question, expected, fields: dict, status: str):
    domain = fields.get("actual_domain", "")
    fill = PatternFill("solid", fgColor=DOMAIN_COLORS.get(domain, DOMAIN_COLORS[""]))

    values = [
        q_id,
        category,
        question,
        expected,
        fields["actual_domain"],
        status,
        fields["model_used"],
        fields["latency_s"],
        "",                       # Relevant? — manual fill
        fields["summary"],
        fields["full_answer"],
        fields["insights"],
        fields["parsed_filters"],
        fields["records_loaded"],
        fields["error"],
    ]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (3, 10, 11, 12)))
        ws.row_dimensions[row_num].height = 60


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model",  default="auto",  help="Model to force (auto/ollama/mistral-small/...)")
    parser.add_argument("--start",  type=int, default=1,  help="Start at question ID (inclusive)")
    parser.add_argument("--end",    type=int, default=50, help="End at question ID (inclusive)")
    parser.add_argument("--out",    default="", help="Output Excel filename (auto-generated if blank)")
    args = parser.parse_args()

    questions = [(q_id, cat, q, exp) for q_id, cat, q, exp in QUESTIONS
                 if args.start <= q_id <= args.end]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(__file__).parent / (args.out or f"benchmark_results_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"
    style_header(ws)

    total = len(questions)
    print(f"\nGovData Benchmark — {total} questions | model={args.model}")
    print(f"Output: {out_path}\n")
    print(f"{'#':>3}  {'Category':<14}  {'Status':<10}  {'Domain':<14}  {'Latency':>8}  Question")
    print("─" * 100)

    for i, (q_id, category, question, expected) in enumerate(questions, start=1):
        row_num = i + 1
        sys.stdout.write(f"{q_id:>3}  {category:<14}  {'submitting':<10}  {'':14}  {'':>8}  {question[:60]}")
        sys.stdout.flush()

        t_start = time.time()
        try:
            sub = submit_query(question, args.model)
            task_id = sub["task_id"]
            task_result = poll_task(task_id)
            elapsed = round(time.time() - t_start, 1)

            state = task_result.get("state", "UNKNOWN")
            fields = extract_fields(task_result)
            # Use wall-clock time if pipeline timing unavailable
            if not fields["latency_s"]:
                fields["latency_s"] = elapsed

            status = "✓ OK" if state == "SUCCESS" else f"✗ {state}"
            print(f"\r{q_id:>3}  {category:<14}  {status:<10}  {fields['actual_domain']:<14}  {elapsed:>7.1f}s  {question[:60]}")

        except Exception as e:
            elapsed = round(time.time() - t_start, 1)
            fields = {k: "" for k in ["actual_domain","model_used","latency_s","summary","full_answer","insights","parsed_filters","records_loaded"]}
            fields["error"] = str(e)
            fields["latency_s"] = elapsed
            status = "✗ ERR"
            print(f"\r{q_id:>3}  {category:<14}  {status:<10}  {'':14}  {elapsed:>7.1f}s  {question[:60]}  [{e}]")

        write_row(ws, row_num, q_id, category, question, expected, fields, status)
        wb.save(out_path)   # save after each question so partial results aren't lost

    print(f"\n{'─' * 100}")
    print(f"Done. Results saved → {out_path}\n")


if __name__ == "__main__":
    main()
